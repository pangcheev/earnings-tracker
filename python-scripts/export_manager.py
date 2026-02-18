"""
Export Manager for Earnings Tracker
Exports data to various formats (CSV, Excel, PDF)
"""

import json
import csv
from datetime import datetime
from typing import List, Dict
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class ExportManager:
    def __init__(self, sessions_data: List[Dict]):
        """
        Initialize export manager
        
        Args:
            sessions_data: List of session dictionaries
        """
        self.sessions = sessions_data
        self.timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    
    def export_to_csv(self, filename: str = None) -> str:
        """
        Export sessions to CSV format
        
        Args:
            filename: Output filename (default: earnings_<timestamp>.csv)
        
        Returns:
            Path to exported file
        """
        if filename is None:
            filename = f"earnings_{self.timestamp}.csv"
        
        fieldnames = [
            'ID', 'Date', 'Location', 'Service Type', 'Duration (min)', 'Rate ($/hr)',
            'Earnings', 'Add-ons', 'Tips', 'Review', 'Client Review'
        ]
        
        rows = []
        
        for session in self.sessions:
            date = session.get('date', '')
            location = session.get('location', '')
            tips = session.get('tips', 0)
            review = session.get('review', '')
            has_review = 'Yes' if session.get('hasClientReview', False) else 'No'
            
            # Add a row for each service
            if session.get('services'):
                for service in session['services']:
                    service_type = service.get('type', '')
                    duration = service.get('duration', 0)
                    rate = service.get('rate', 0)
                    earnings = (rate / 60) * duration
                    
                    addons_str = ', '.join([a.get('name', '') for a in session.get('addOns', [])])
                    addons_total = sum([a.get('price', 0) for a in session.get('addOns', [])])
                    
                    rows.append({
                        'ID': session.get('id', ''),
                        'Date': date,
                        'Location': location,
                        'Service Type': service_type,
                        'Duration (min)': duration,
                        'Rate ($/hr)': f"${rate:.2f}",
                        'Earnings': f"${earnings:.2f}",
                        'Add-ons': addons_str if addons_str else 'None',
                        'Tips': f"${tips:.2f}",
                        'Review': review,
                        'Client Review': has_review,
                    })
            else:
                # No services
                addons_str = ', '.join([a.get('name', '') for a in session.get('addOns', [])])
                rows.append({
                    'ID': session.get('id', ''),
                    'Date': date,
                    'Location': location,
                    'Service Type': 'N/A',
                    'Duration (min)': 'N/A',
                    'Rate ($/hr)': 'N/A',
                    'Earnings': '$0.00',
                    'Add-ons': addons_str if addons_str else 'None',
                    'Tips': f"${tips:.2f}",
                    'Review': review,
                    'Client Review': has_review,
                })
        
        # Write CSV
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        
        print(f"✓ CSV export saved to: {filename}")
        return filename
    
    def export_to_excel(self, filename: str = None) -> str:
        """
        Export sessions to Excel format
        
        Args:
            filename: Output filename (default: earnings_<timestamp>.xlsx)
        
        Returns:
            Path to exported file
        """
        if not HAS_OPENPYXL:
            print("❌ openpyxl not installed. Install with: pip install openpyxl")
            return None
        
        if filename is None:
            filename = f"earnings_{self.timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sessions"
        
        # Headers
        headers = ['Date', 'Location', 'Service Type', 'Duration', 'Rate', 'Earnings', 'Add-ons', 'Tips', 'Total']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        
        # Add data
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for session in self.sessions:
            date = session.get('date', '')
            location = session.get('location', '')
            tips = float(session.get('tips', 0))
            
            services = session.get('services', [])
            addons = session.get('addOns', [])
            addons_total = sum(float(a.get('price', 0)) for a in addons)
            
            for service in services:
                service_type = service.get('type', '')
                duration = float(service.get('duration', 0))
                rate = float(service.get('rate', 0))
                earnings = (rate / 60) * duration
                
                addons_str = ', '.join([a.get('name', '') for a in addons])
                total = earnings + addons_total + tips
                
                row = [date, location, service_type, duration, f"${rate:.2f}", f"${earnings:.2f}", addons_str, f"${tips:.2f}", f"${total:.2f}"]
                ws.append(row)
                
                for cell in ws[ws.max_row]:
                    cell.border = border
        
        wb.save(filename)
        print(f"✓ Excel export saved to: {filename}")
        return filename
    
    def export_to_pdf(self, filename: str = None) -> str:
        """
        Export sessions to PDF format
        
        Args:
            filename: Output filename (default: earnings_<timestamp>.pdf)
        
        Returns:
            Path to exported file
        """
        if not HAS_REPORTLAB:
            print("❌ reportlab not installed. Install with: pip install reportlab")
            return None
        
        if filename is None:
            filename = f"earnings_{self.timestamp}.pdf"
        
        # Create PDF
        pdf = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Custom title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Build story
        story = []
        
        # Title
        story.append(Paragraph("Earnings Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary table
        summary_data = [['Metric', 'Value']]
        total_earnings = sum(
            sum((s.get('rate', 0) / 60) * s.get('duration', 0) for s in session.get('services', []))
            for session in self.sessions
        )
        total_tips = sum(session.get('tips', 0) for session in self.sessions)
        
        summary_data.append(['Total Sessions', str(len(self.sessions))])
        summary_data.append(['Total Earnings', f"${total_earnings:.2f}"])
        summary_data.append(['Total Tips', f"${total_tips:.2f}"])
        summary_data.append(['Gross Total', f"${total_earnings + total_tips:.2f}"])
        
        table = Table(summary_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.3*inch))
        
        # Sessions table
        story.append(Paragraph("Session Details", styles['Heading2']))
        
        sessions_data = [['Date', 'Location', 'Services', 'Tips', 'Total']]
        
        for session in sorted(self.sessions, key=lambda x: x.get('date', ''), reverse=True):
            date = session.get('date', '')
            location = session.get('location', '')
            services_str = ', '.join([s.get('type', '') for s in session.get('services', [])])
            tips = session.get('tips', 0)
            
            earnings = sum((s.get('rate', 0) / 60) * s.get('duration', 0) for s in session.get('services', []))
            addons = sum(a.get('price', 0) for a in session.get('addOns', []))
            total = earnings + addons + tips
            
            sessions_data.append([date, location, services_str, f"${tips:.2f}", f"${total:.2f}"])
        
        sessions_table = Table(sessions_data)
        sessions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        story.append(sessions_table)
        
        # Build PDF
        pdf.build(story)
        print(f"✓ PDF export saved to: {filename}")
        return filename

def load_sessions_from_backup(backup_file: str) -> List[Dict]:
    """Load sessions from a JSON backup file"""
    try:
        with open(backup_file, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Backup file not found: {backup_file}")
        return []

def main():
    # Example usage
    backup_file = "earnings-backup-2026-02-17.json"
    
    sessions = load_sessions_from_backup(backup_file)
    
    if sessions:
        exporter = ExportManager(sessions)
        
        # Export to CSV
        exporter.export_to_csv()
        
        # Export to Excel (if available)
        exporter.export_to_excel()
        
        # Export to PDF (if available)
        exporter.export_to_pdf()

if __name__ == '__main__':
    main()
